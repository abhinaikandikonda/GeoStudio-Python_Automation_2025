import pandas as pd                    #for excel works
import xml.etree.ElementTree as ET         #for xml reading





# four steps need to falllow by the user
# step 1 name of the xml file 
# step 2 ^^^(number of files + 1 )^^^ in the range
# step 3 number of values that is number of files
# step 4 name of the excel sheet for input


######################################## things to change by the user ################################################################################################


xml_file_path = "D:/Abinai_Mtech/acess_xml/for_12m35 -1st.xml"


file_names1 = [f'D:/Abinai_Mtech/xmlfiles/file_{i}.xml' for i in range(1,1201)] # it will store the genereated xml file here(change acccordingly)

num_values = 1200 # Example: assign one value less than than the above  range length 


parameters = pd.read_excel("D:/Abinai_Mtech/excel_sheets_c/input_for_varying_cs9.xlsx") # this is for reading the excel data and store the excel file data(change acccordingly where the excel file is located)


####################################   the end for the user #####################################################################################################







# Parse the XML file using ElementTree
tree = ET.parse(xml_file_path)


root = tree.getroot()

# file_names1 = [f'C:/Users/abhia/OneDrive/Documents/phyton foundation course/comming4/file_8{i}.xml' for i in range(41, 51)]


################## Define row ranges################
# print(parameters['Soil Type'])
# print(parameters['Soil Type'].str.contains('material-', na=False))
# print(parameters[parameters['Soil Type'].str.contains('material-', na=False)])

material_indices = parameters[parameters['Soil Type'].str.contains('material-', na=False)].index

material_sections = {}
print(material_indices)

aRes1=[]
nRes1=[]
aRes2=[]
nRes2=[]
aRes3=[]
nRes3=[]
aRes4=[]
nRes4=[]
thRes1 = []
thRes2 = []
thRes3 = []
thRes4 = []
tsRes1 = []
tsRes2= []
tsRes3 = []
tsRes4= []
mRes1=[]
mRes2=[]
mRes3=[]
mRes4=[]
# Loop through each material section
for i in range(len(material_indices)):
    # Start right after the 'material-X' row
    start_index = material_indices[i] + 1
    
    # Determine the end index (start of the next material or end of the DataFrame)
    if i < len(material_indices) - 1:
        end_index = material_indices[i + 1]
    else:
        end_index = len(parameters)
    
    # Get the rows for the current material section
    material_rows = parameters.iloc[start_index:end_index]
    
    # Store the section in a dictionary
    material_sections[f'material_{i+1}'] = material_rows
    
# print(material_sections)
# Now you can access each section like this:
# material_sections['material_1'], material_sections['material_2'], etc.

rows_1 = material_sections['material_1']
rows_2 =   material_sections['material_2']
rows_3 = material_sections['material_3']
rows_4 =material_sections['material_4']
# rows_5 = material_sections['material_5']
# rows_6 = material_sections['material_6']
print(rows_1) 
print(rows_2)
print(rows_3)
print(rows_4)
# print(rows_5)
# print(rows_6)





# rows_1 = parameters.iloc[1:11] 
# rows_2 = parameters.iloc[12:22]  
# rows_3 = parameters.iloc[23:33]
# rows_4 = parameters.iloc[34:44]
# rows_5 = parameters.iloc[45:55]
# rows_6 = parameters.iloc[56:66]
# print(rows_1) 
# print("abhi")
# print(rows_2) 
# print('by abhi')
# print(rows_3)
# print(rows_4)
# print(rows_5)
# print(rows_6)


# Find the 'Materials' element
soil_element = root.find('Materials')



kfn_elements = root.findall(".//Functions/Material/Hydraulic/KFns/KFn")  # Find all KFn elements
# first_kfn = kfn_elements[0]  # Access the first KFn (index 0)
#kfn_elements[i][1]== "hc dyke ash"

# # Access the first Point within this KFn and get the Y attribute
# first_point = first_kfn.find(".//Points/Point[@X='0']")  # Find the Point where X="0"
# if first_point is not None:
#     y_value = first_point.get("Y")  # Get the Y attribute
#     print("Y value for the first Point in the first KFn:", y_value)
# else:
#     print("Point with X='0' not found.")


for i in range(0, 10):
    if soil_element[i][2].text == "starter dyke soil":
        unitweight1 = soil_element[i][5][4]
        print("your calculation")
        
        cohesion1 = soil_element[i][5][5]
        
        angle1 = soil_element[i][5][6]
        print("for starter")
        print(angle1.text)
        print(unitweight1.text)
        print(cohesion1.text)
    elif soil_element[i][2].text == "foundation":
        unitweight2 = soil_element[i][5][4]
        cohesion2 = soil_element[i][5][5]
        angle2 = soil_element[i][5][6]
        print("for foundation")
        print(angle2.text)
        print(unitweight2.text)
        print(cohesion2.text)        
    elif soil_element[i][2].text=="dyke ash" :
        unitweight3 = soil_element[i][5][4]
        cohesion3 = soil_element[i][5][5]
        angle3 = soil_element[i][5][6]
        print("forr dyke")
        print(angle3.text)
        print(unitweight3.text)
        print(cohesion3.text)           
        
        #permeability3= for_permeability[i][3][1]
    elif soil_element[i][2].text=="pond ash":
        unitweight4 = soil_element[i][5][4]
        cohesion4 = soil_element[i][5][5]
        angle4 = soil_element[i][5][6]
        print("for pond ")
        print(angle4.text)
        print(unitweight4.text)
        print(cohesion4.text)   
    # elif soil_element[i][2].text=="foundation":
    #     unitweight5 = soil_element[i][5][4]
    #     cohesion5 = soil_element[i][5][5]
    #     angle5 = soil_element[i][5][6]
    # elif soil_element[i][2].text=="starter dyke (below phreatic line)":
    #     unitweight6 = soil_element[i][5][4]
    #     cohesion6 = soil_element[i][5][5]
    #     angle6 = soil_element[i][5][6]        end

print(len(kfn_elements))

# for i in range(len(kfn_elements)):  # Dynamically iterate through all elements
#     if kfn_elements[i][1].text == "hc dyke ash":
#         print(i)
#         permeability_dyke_ash = kfn_elements[i][4][0]

#         print("Type of permeability_dyke_ash:", (permeability_dyke_ash.get("Y")))

#     elif kfn_elements[i][1].text== "hc for pond ash": 
#         print(i)
#         permeability_pond_Ash = kfn_elements[i][3][0] 
#         print("Type of permeability_dyke_ash:", (permeability_pond_Ash.get("Y")))
#     elif kfn_elements[i][1].text == "hc for starter dyke( above)":

#         print(i)
#         permeability_starter_dyke = kfn_elements[i][3][0] 
#         print("Type of permeability_dyke_ash:", (permeability_starter_dyke.get("Y")))




# for i in range(0,1):
#     if kfn_elements[i][1]== "hc dyke ash" :
#         permeability_dyke_ash = kfn_elements[i][3][0] 
#         print(permeability_dyke_ash.get("Y"))
        # cohesion1 = soil_element[i][5][5]
        # angle1 = soil_element[i][5][6]

    #elif kfn_elements[i][1]== "hc for pond ash":
        # unitweight2 = soil_element[i][5][4]
        # cohesion2 = soil_element[i][5][5]
        # angle2 = soil_element[i][5][6]
    #elif kfn_elements[i][1]== "hc for starter dyke( above)" :
        # unitweight3 = soil_element[i][5][4]
        # cohesion3 = soil_element[i][5][5]
        # angle3 = soil_element[i][5][6]
        

    
    # elif soil_element[i][2].text=="foundation":
    #     unitweight5 = soil_element[i][5][4]
    #     cohesion5 = soil_element[i][5][5]
    #     angle5 = soil_element[i][5][6]
    # elif soil_element[i][2].text=="starter dyke (below phreatic line)":
    #     unitweight6 = soil_element[i][5][4]
    #     cohesion6 = soil_element[i][5][5]
    #     angle6 = soil_element[i][5][6]        


# permeability_dyke_ash = kfn_elements[0][3][0] 

# print(permeability_dyke_ash.get("Y"))




  
print('abhitest')


# Iterate over the rows in the DataFrame and update the XML elements
for idx, row in parameters.iterrows():
    if idx < len(file_names1):  # First 5 rows # here we need to run for 10 xml file so provide 10 if it increases then increase the number or else give length of something which i will see later
        i=1
        unitweight1.text = str(rows_1.iloc[idx]['Unit Weight (kN/m³)'])
        cohesion1.text = str(rows_1.iloc[idx]['Cohesion (kPa)'])
        angle1.text = str(rows_1.iloc[idx]['Angle of Internal Friction (°)'])
        K_sat_starter =float(rows_1.iloc[idx]['permeability'])
        # permeability_starter_dyke.set("Y", str(rows_1.iloc[idx]['permeability']))
        
        print(f'Updated starter dyke soil values: unitweight={unitweight1.text}, cohesion={cohesion1.text}, angle of internal friction={angle1.text},')
    # elif 5 <= idx < 10:  # Next 5 rows     #material2
        unitweight2.text = str(rows_2.iloc[idx]['Unit Weight (kN/m³)'])
        print('abhi')
        #print(unitweight2.text)
        
        cohesion2.text = str(rows_2.iloc[idx]['Cohesion (kPa)'])
        angle2.text = str(rows_2.iloc[idx]['Angle of Internal Friction (°)'])
        K_sat_found =float(rows_2.iloc[idx]['permeability'])
        print(f'Updated foundation values: unitweight={unitweight2.text}, cohesion={cohesion2.text},angle of internal friction= {angle2.text}')
        #material3
        unitweight3.text = str(rows_3.iloc[idx]['Unit Weight (kN/m³)'])
        cohesion3.text = str(rows_3.iloc[idx]['Cohesion (kPa)'])
        angle3.text = str(rows_3.iloc[idx]['Angle of Internal Friction (°)'])
        # permeability_dyke_ash.set("Y", str(rows_3.iloc[idx]['permeability']))
        K_sat_dyke = float(rows_3.iloc[idx]['permeability'])
        
        print(f'Updated dyke ash values: unitweight={unitweight3.text}, cohesion={cohesion3.text}, angle of internal friction={angle3.text}')
        #material4
        
        unitweight4.text = str(rows_4.iloc[idx]['Unit Weight (kN/m³)'])
        cohesion4.text = str(rows_4.iloc[idx]['Cohesion (kPa)'])
        angle4.text = str(rows_4.iloc[idx]['Angle of Internal Friction (°)'])
        K_sat_pond = float(rows_4.iloc[idx]['permeability'])
        # permeability_pond_Ash.set("Y", str(rows_4.iloc[idx]['permeability']))
        print(f'Updated pond ash values: unitweight={unitweight4.text}, coheison={cohesion4.text},angle of internal friction {angle4.text}')





        import numpy as np
        import matplotlib.pyplot as plt
        import pandas as pd  # If rows_2 is from foundation DataFrame

        # Define the Fredlund and Xing (1994) model for volumetric water content
        def fredlund_xing(suction, theta_s, a, n, m, psi_r=1000):
            """
            Calculate volumetric water content based on the Fredlund and Xing (1994) model.

            Parameters:
            - suction: Array of suction values (kPa)
            - theta_s: Saturated volumetric water content
            - a: Parameter related to the air entry value
            - n: Parameter controlling the slope of the SWCC
            - m: Parameter controlling the curvature of the SWCC
            - psi_r: Residual suction value for correction factor C(psi) (default is 1000 kPa)

            Returns:
            - Array of volumetric water content values corresponding to each suction value
            """
            # Compute the correction factor
            correction_factor = 1 - (np.log(1 + (suction / psi_r)) / np.log(1 + (10**6 / psi_r)))
            
            # Compute the main SWCC equation
            theta = theta_s *correction_factor * (1 / np.log(np.e + (suction / a)**n))**m
            return theta

        # Define the hydraulic conductivity function based on Fredlund & Xing model
        def hydraulic_conductivity_fredlund_xing(theta, theta_s, theta_r, K_sat, L=0.5):
            """
            Calculate hydraulic conductivity based on volumetric water content 
            using the Fredlund & Xing (1994) model.

            Parameters:
            - theta: Volumetric water content at given suction
            - theta_s: Saturated volumetric water content
            - theta_r: Residual volumetric water content
            - K_sat: Saturated hydraulic conductivity (m/s)
            - L: Empirical pore connectivity parameter (default is 0.5)

            Returns:
            - Array of hydraulic conductivity values corresponding to each water content
            """
            # Ensure no division by zero when (theta_s - theta_r) is very small
            if np.isclose(theta_s, theta_r):
                raise ValueError("Saturated and residual water content should not be identical.")

            return K_sat * ((theta - theta_r) / (theta_s - theta_r)) ** L

        # Define parameters for SWCC and hydraulic conductivity (example values)
        theta_s = 0.5    # Saturated volumetric water content
        theta_r = 0.04   # Residual volumetric water content
        a = float(rows_2.iloc[idx]['a'])  # Parameter related to the air entry value
        n = float(rows_2.iloc[idx]['n'])  # Parameter controlling the slope of the SWCC
        m = float(rows_2.iloc[idx]['m'])  # Parameter controlling the curvature of the 
        aRes2.append(a)
        nRes2.append(n)
        mRes2.append(m)
        thRes2.append(theta_r)
        tsRes2.append(theta_s)
        
        psi_r = 1500     # Residual suction (default or typical value, can be adjusted)
        K_sat_found   # Example value for saturated hydraulic conductivity (m/s)
        L = 0.5          # Pore connectivity parameter

        # Define a range of suction values (e.g., from 1 kPa to 1,000,000 kPa)
        suction_values = [
            0, 0.01, 0.018329807, 0.033598183, 0.061584821, 0.11288379, 
            0.20691381, 0.37926902, 0.6951928, 1.274275, 2.3357215, 
            4.2813324, 7.8475997, 14.384499, 26.366509, 48.329302, 
            88.586679, 162.37767, 297.63514, 545.55948, 1000
        ]
        #uction_values = np.logspace(0, 6, 500)  # Logarithmic scale from 1 to 1,000,000 kPa

        # Calculate volumetric water content for each suction value
        water_content_values = fredlund_xing(np.array(suction_values), theta_s, a, n, m, psi_r)

        # Calculate hydraulic conductivity for each volumetric water content
        hydraulic_conductivity_values = hydraulic_conductivity_fredlund_xing(water_content_values, theta_s, theta_r, K_sat_found, L)

        # Store the results in a structured array or DataFrame
        import math
        formatted_data_found_vwc = [
            (float(suction), float(water_content)) 
            for suction, water_content in zip(suction_values, water_content_values)
            if not math.isnan(float(water_content))  # Filter out nan values
            #if not math.isnan(float(water_content)) and float(water_content) != 0  # Filter out NaN and zero values
        ]
        
        print(formatted_data_found_vwc)


        data_points = pd.DataFrame({
            "Suction (kPa)": suction_values,
            "Volumetric Water Content": water_content_values,
            "Hydraulic Conductivity (m/s)": hydraulic_conductivity_values
        })
        print(data_points)
        
        import math
        formatted_data_found_hc = [
            (float(suction), float(hc)) 
            for suction, hc in zip(suction_values, hydraulic_conductivity_values)
            if not math.isnan(float(hc))  # Filter out nan values
            #if not math.isnan(float(hc)) and float(hc) != 0  # Filter out NaN and zero values
        ]
        print(formatted_data_found_hc)
        
        # Plot Volumetric Water Content vs. Suction
        # plt.figure(figsize=(12, 6))
        # plt.subplot(1, 2, 1)
        # plt.plot(suction_values, water_content_values, label="Volumetric Water Content", color="blue")
        # plt.xscale('log')
        # plt.xlabel("Suction (kPa)")
        # plt.ylabel("Volumetric Water Content ")
        # plt.title("Soil-Water Characteristic Curve for found")
        # plt.legend()
        # plt.grid(True)

        # # Plot Hydraulic Conductivity vs. Suction
        # plt.subplot(1, 2, 2)
        # plt.plot(suction_values, hydraulic_conductivity_values, label="Hydraulic Conductivity", color="red")
        # plt.xscale('log')
        # plt.yscale('log')
        # plt.xlabel("Suction (kPa)")
        # plt.ylabel("Hydraulic Conductivity (m/s)")    
        # plt.title("Hydraulic Conductivity Curve")     
        # plt.legend()                               
        # plt.grid(True)                               
                                      
        # # Show plots          
        # plt.tight_layout()
        # plt.show()









        import numpy as np
        import matplotlib.pyplot as plt
        import pandas as pd  # If rows_3 is from a DataFrame for dyke ash

        # Define the Fredlund and Xing (1994) model for volumetric water content
        def fredlund_xing(suction, theta_s, a, n, m, psi_r=1000):
            """
            Calculate volumetric water content based on the Fredlund and Xing (1994) model.

            Parameters:
            - suction: Array of suction values (kPa)
            - theta_s: Saturated volumetric water content
            - a: Parameter related to the air entry value
            - n: Parameter controlling the slope of the SWCC
            - m: Parameter controlling the curvature of the SWCC
            - psi_r: Residual suction value for correction factor C(psi) (default is 1000 kPa)

            Returns:
            - Array of volumetric water content values corresponding to each suction value
            """
            # Compute the correction factor
            correction_factor = 1 - (np.log(1 + (suction / psi_r)) / np.log(1 + (10**6 / psi_r)))
            
            # Compute the main SWCC equation
            theta = theta_s *correction_factor * (1 / np.log(np.e + (suction / a)**n))**m
            return theta

        # Define the hydraulic conductivity function based on Fredlund & Xing model
        def hydraulic_conductivity_fredlund_xing(theta, theta_s, theta_r, K_sat, L=0.5):
            """
            Calculate hydraulic conductivity based on volumetric water content 
            using the Fredlund & Xing (1994) model.

            Parameters:
            - theta: Volumetric water content at given suction
            - theta_s: Saturated volumetric water content
            - theta_r: Residual volumetric water content
            - K_sat: Saturated hydraulic conductivity (m/s)
            - L: Empirical pore connectivity parameter (default is 0.5)

            Returns:
            - Array of hydraulic conductivity values corresponding to each water content
            """
            # Ensure no division by zero when (theta_s - theta_r) is very small
            if np.isclose(theta_s, theta_r):
                raise ValueError("Saturated and residual water content should not be identical.")

            return K_sat * ((theta - theta_r) / (theta_s - theta_r)) ** L

        # Define parameters for SWCC and hydraulic conductivity (example values)
        theta_s = 0.4   # Saturated volumetric water content
        theta_r = 0.03   # Residual volumetric water content
        a = float(rows_3.iloc[idx]['a']) 
        n = float(rows_3.iloc[idx]['n'])  # Parameter controlling the slope of the SWCC
        m = float(rows_3.iloc[idx]['m'])  # Parameter controlling the curvature of the SWCC
        aRes3.append(a)
        nRes3.append(n)
        mRes3.append(m)
        thRes3.append(theta_r)
        tsRes3.append(theta_s)
        psi_r = 1500     # Residual suction (default or typical value, can be adjusted)
        K_sat_dyke   # Example value for saturated hydraulic conductivity (m/s)
        L = 0.5          # Pore connectivity parameter

        # Define a range of suction values (e.g., from 1 kPa to 1,000,000 kPa)
        suction_values = [
            0, 0.01, 0.018329807, 0.033598183, 0.061584821, 0.11288379, 
            0.20691381, 0.37926902, 0.6951928, 1.274275, 2.3357215, 
            4.2813324, 7.8475997, 14.384499, 26.366509, 48.329302, 
            88.586679, 162.37767, 297.63514, 545.55948, 1000
        ]
        #suction_values = np.logspace(0, 6, 500)  # Logarithmic scale from 1 to 1,000,000 kPa

        # Calculate volumetric water content for each suction value
        water_content_values = fredlund_xing(np.array(suction_values), theta_s, a, n, m, psi_r)

        # Calculate hydraulic conductivity for each volumetric water content
        hydraulic_conductivity_values = hydraulic_conductivity_fredlund_xing(water_content_values, theta_s, theta_r, K_sat_dyke, L)

        # Store the results in a structured array or DataFrame
        import math
        formatted_data_dyke_vwc = [
            (float(suction), float(water_content)) 
            for suction, water_content in zip(suction_values, water_content_values)
            if not math.isnan(float(water_content))  # Filter out nan values
            #if not math.isnan(float(water_content)) and float(water_content) != 0  # Filter out NaN and zero values
        ]

        print(formatted_data_dyke_vwc)

        
        import math
        formatted_data_dyke_hc = [
            (float(suction), float(hc)) 
            for suction, hc in zip(suction_values, hydraulic_conductivity_values)
            if not math.isnan(float(hc))  # Filter out nan values
            #if not math.isnan(float(hc)) and float(hc) != 0  # Filter out NaN and zero values
            
        ]
        print(formatted_data_dyke_hc)
        
        data_points = pd.DataFrame({
            "Suction (kPa)": suction_values,
            "Volumetric Water Content": water_content_values,
            "Hydraulic Conductivity (m/s)": hydraulic_conductivity_values
        })
        print(data_points)
        # Plot Volumetric Water Content vs. Suction
        # plt.figure(figsize=(12, 6))
        # plt.subplot(1, 2, 1)
        # plt.plot(suction_values, water_content_values, label="Volumetric Water Content", color="blue")
        # plt.xscale('log')
        # plt.xlabel("Suction (kPa)")
        # plt.ylabel("Volumetric Water Content ")
        # plt.title("Soil-Water Characteristic Curve for dyke ")
        # plt.legend()
        # plt.grid(True)

        # # Plot Hydraulic Conductivity vs. Suction
        # plt.subplot(1, 2, 2)
        # plt.plot(suction_values, hydraulic_conductivity_values, label="Hydraulic Conductivity", color="red")
        # plt.xscale('log')
        # plt.yscale('log')
        # plt.xlabel("Suction (kPa)")
        # plt.ylabel("Hydraulic Conductivity (m/s)")    
        # plt.title("Hydraulic Conductivity Curve")     
        # plt.legend()                               
        # plt.grid(True)                               
                                      
        # # Show plots          
        # plt.tight_layout()
        # plt.show()










        import numpy as np
        import matplotlib.pyplot as plt
        import pandas as pd  # If rows_4 is from a DataFrame for pond ash

        # Define the Fredlund and Xing (1994) model for volumetric water content
        def fredlund_xing(suction, theta_s, a, n, m, psi_r=1000):
            """
            Calculate volumetric water content based on the Fredlund and Xing (1994) model.

            Parameters:
            - suction: Array of suction values (kPa)
            - theta_s: Saturated volumetric water content
            - a: Parameter related to the air entry value
            - n: Parameter controlling the slope of the SWCC
            - m: Parameter controlling the curvature of the SWCC
            - psi_r: Residual suction value for correction factor C(psi) (default is 1000 kPa)

            Returns:
            - Array of volumetric water content values corresponding to each suction value
            """
            # Compute the correction factor
            correction_factor = 1 - (np.log(1 + (suction / psi_r)) / np.log(1 + (10**6 / psi_r)))
            
            # Compute the main SWCC equation
            theta = theta_s *correction_factor * (1 / np.log(np.e + (suction / a)**n))**m
            return theta

        # Define the hydraulic conductivity function based on Fredlund & Xing model
        def hydraulic_conductivity_fredlund_xing(theta, theta_s, theta_r, K_sat, L=0.5):
            """
            Calculate hydraulic conductivity based on volumetric water content 
            using the Fredlund & Xing (1994) model.

            Parameters:
            - theta: Volumetric water content at given suction
            - theta_s: Saturated volumetric water content
            - theta_r: Residual volumetric water content
            - K_sat: Saturated hydraulic conductivity (m/s)
            - L: Empirical pore connectivity parameter (default is 0.5)

            Returns:
            - Array of hydraulic conductivity values corresponding to each water content
            """
            # Ensure no division by zero when (theta_s - theta_r) is very small
            if np.isclose(theta_s, theta_r):
                raise ValueError("Saturated and residual water content should not be identical.")

            return K_sat * ((theta - theta_r) / (theta_s - theta_r)) ** L

        # Define parameters for SWCC and hydraulic conductivity (example values)
        theta_s = 0.3    # Saturated volumetric water content
        theta_r = 0.02   # Residual volumetric water content
        a = float(rows_4.iloc[idx]['a']) 
        n = float(rows_4.iloc[idx]['n'])  # Parameter controlling the slope of the SWCC
        m = float(rows_4.iloc[idx]['m'])  # Parameter controlling the curvature of the SWCC
        aRes4.append(a)
        nRes4.append(n)
        mRes4.append(m)
        thRes4.append(theta_r)
        tsRes4.append(theta_s)
        psi_r = 1500     # Residual suction (default or typical value, can be adjusted)
        K_sat_pond   # Example value for saturated hydraulic conductivity (m/s)
        L = 0.5          # Pore connectivity parameter

        # Define a range of suction values (e.g., from 1 kPa to 1,000,000 kPa)
        
        suction_values = [
            0, 0.01, 0.018329807, 0.033598183, 0.061584821, 0.11288379, 
            0.20691381, 0.37926902, 0.6951928, 1.274275, 2.3357215, 
            4.2813324, 7.8475997, 14.384499, 26.366509, 48.329302, 
            88.586679, 162.37767, 297.63514, 545.55948, 1000
        ]
        #suction_values = np.logspace(0, 6, 500)  # Logarithmic scale from 1 to 1,000,000 kPa

        # Calculate volumetric water content for each suction value
        water_content_values = fredlund_xing(np.array(suction_values), theta_s, a, n, m, psi_r)

        # Calculate hydraulic conductivity for each volumetric water content
        hydraulic_conductivity_values = hydraulic_conductivity_fredlund_xing(water_content_values, theta_s, theta_r, K_sat_pond, L)

        # Store the results in a structured array or DataFrame
        import math
        formatted_data_pond_vwc = [
            (float(suction), float(water_content)) 
            for suction, water_content in zip(suction_values, water_content_values)
            if not math.isnan(float(water_content))  # Filter out nan values
            #if not math.isnan(float(water_content)) and float(water_content) != 0  # Filter out NaN and zero values
        ]

        print(formatted_data_pond_vwc)

        
        import math
        formatted_data_pond_hc = [
            (float(suction), float(hc)) 
            for suction, hc in zip(suction_values, hydraulic_conductivity_values)
            if not math.isnan(float(hc))  # Filter out nan values
            #if not math.isnan(float(hc)) and float(hc) != 0
        ]
        print(formatted_data_pond_hc)
        
        data_points = pd.DataFrame({
            "Suction (kPa)": suction_values,
            "Volumetric Water Content": water_content_values,
            "Hydraulic Conductivity (m/s)": hydraulic_conductivity_values
        })
        print(data_points)
        # Plot Volumetric Water Content vs. Suction
        # plt.figure(figsize=(12, 6))
        # plt.subplot(1, 2, 1)
        # plt.plot(suction_values, water_content_values, label="Volumetric Water Content", color="blue")
        # plt.xscale('log')
        # plt.xlabel("Suction (kPa)")
        # plt.ylabel("Volumetric Water Content ")
        # plt.title("Soil-Water Characteristic Curve for dyke ")
        # plt.legend()
        # plt.grid(True)

        # # Plot Hydraulic Conductivity vs. Suction
        # plt.subplot(1, 2, 2)
        # plt.plot(suction_values, hydraulic_conductivity_values, label="Hydraulic Conductivity", color="red")
        # plt.xscale('log')
        # plt.yscale('log')
        # plt.xlabel("Suction (kPa)")
        # plt.ylabel("Hydraulic Conductivity (m/s)")    
        # plt.title("Hydraulic Conductivity Curve")     
        # plt.legend()                               
        # plt.grid(True)                               
                                      
        # # Show plots          
        # plt.tight_layout()
        # plt.show()









        import numpy as np
        import matplotlib.pyplot as plt
        import pandas as pd  # If rows_1 is from a DataFrame for starter ash

        # Define the Fredlund and Xing (1994) model for volumetric water content
        def fredlund_xing(suction, theta_s, a, n, m, psi_r=1000):
            """
            Calculate volumetric water content based on the Fredlund and Xing (1994) model.

            Parameters:
            - suction: Array of suction values (kPa)
            - theta_s: Saturated volumetric water content
            - a: Parameter related to the air entry value
            - n: Parameter controlling the slope of the SWCC
            - m: Parameter controlling the curvature of the SWCC
            - psi_r: Residual suction value for correction factor C(psi) (default is 1000 kPa)

            Returns:
            - Array of volumetric water content values corresponding to each suction value
            """
            # Compute the correction factor
            correction_factor = 1 - (np.log(1 + (suction / psi_r)) / np.log(1 + (10**6 / psi_r)))
            
            # Compute the main SWCC equation
            theta = theta_s *correction_factor * (1 / np.log(np.e + (suction / a)**n))**m
            return theta

        # Define the hydraulic conductivity function based on Fredlund & Xing model
        def hydraulic_conductivity_fredlund_xing(theta, theta_s, theta_r, K_sat, L=0.5):
            """
            Calculate hydraulic conductivity based on volumetric water content 
            using the Fredlund & Xing (1994) model.

            Parameters:
            - theta: Volumetric water content at given suction
            - theta_s: Saturated volumetric water content
            - theta_r: Residual volumetric water content
            - K_sat: Saturated hydraulic conductivity (m/s)
            - L: Empirical pore connectivity parameter (default is 0.5)

            Returns:
            - Array of hydraulic conductivity values corresponding to each water content
            """
            # Ensure no division by zero when (theta_s - theta_r) is very small
            if np.isclose(theta_s, theta_r):
                raise ValueError("Saturated and residual water content should not be identical.")

            return K_sat * ((theta - theta_r) / (theta_s - theta_r)) ** L

        # Define parameters for SWCC and hydraulic conductivity (example values)
        theta_s = 0.1    # Saturated volumetric water content
        theta_r = 0.01   # Residual volumetric water content
        a = float(rows_1.iloc[idx]['a']) 
        n = float(rows_1.iloc[idx]['n'])  # Parameter controlling the slope of the SWCC
        m = float(rows_1.iloc[idx]['m'])
        aRes1.append(a)
        nRes1.append(n)
        mRes1.append(m)
        thRes1.append(theta_r)
        tsRes1.append(theta_s)  # Parameter controlling the curvature of the SWCC
        psi_r = 1500     # Residual suction (default or typical value, can be adjusted)
         # Example value for saturated hydraulic conductivity (m/s)
        L = 0.5          # Pore connectivity parameter

        # Define a range of suction values (e.g., from 1 kPa to 1,000,000 kPa)
        suction_values = [
            0, 0.01, 0.018329807, 0.033598183, 0.061584821, 0.11288379, 
            0.20691381, 0.37926902, 0.6951928, 1.274275, 2.3357215, 
            4.2813324, 7.8475997, 14.384499, 26.366509, 48.329302, 
            88.586679, 162.37767, 297.63514, 545.55948, 1000
        ]
        #suction_values = np.logspace(0, 6, 500)  # Logarithmic scale from 1 to 1,000,000 kPa

        # Calculate volumetric water content for each suction value
        water_content_values = fredlund_xing(np.array(suction_values), theta_s, a, n, m, psi_r)

        # Calculate hydraulic conductivity for each volumetric water content
        hydraulic_conductivity_values = hydraulic_conductivity_fredlund_xing(water_content_values, theta_s, theta_r, K_sat_starter, L)

        # Store the results in a structured array or DataFrame
        import math
        formatted_data_starter_vwc = [
            (float(suction), float(water_content)) 
            for suction, water_content in zip(suction_values, water_content_values)
            if not math.isnan(float(water_content))  # Filter out nan values
            #if not math.isnan(float(water_content)) and float(water_content) != 0  # Filter out NaN and zero values
        ]

        print(formatted_data_starter_vwc)

        
        import math
        formatted_data_starter_hc = [
            (float(suction), float(hc)) 
            for suction, hc in zip(suction_values, hydraulic_conductivity_values)
            if not math.isnan(float(hc))  # Filter out nan values
            #if not math.isnan(float(hc)) and float(hc) != 0  # Filter out NaN and zero values
        ]
        print(formatted_data_starter_hc)
        
        data_points = pd.DataFrame({
            "Suction (kPa)": suction_values,
            "Volumetric Water Content": water_content_values,
            "Hydraulic Conductivity (m/s)": hydraulic_conductivity_values
        })
        print(data_points)
        # # Plot Volumetric Water Content vs. Suction
        # plt.figure(figsize=(12, 6))
        # plt.subplot(1, 2, 1)
        # plt.plot(suction_values, water_content_values, label="Volumetric Water Content", color="blue")
        # plt.xscale('log')
        # plt.xlabel("Suction (kPa)")
        # plt.ylabel("Volumetric Water Content ")
        # plt.title("Soil-Water Characteristic Curve for starter ")
        # plt.legend()
        # plt.grid(True)

        # # Plot Hydraulic Conductivity vs. Suction
        # plt.subplot(1, 2, 2)
        # plt.plot(suction_values, hydraulic_conductivity_values, label="Hydraulic Conductivity", color="red")
        # plt.xscale('log')
        # plt.yscale('log')
        # plt.xlabel("Suction (kPa)")
        # plt.ylabel("Hydraulic Conductivity (m/s)")    
        # plt.title("Hydraulic Conductivity Curve")     
        # plt.legend()                               
        # plt.grid(True)                               
                                      
        # # Show plots          
        # plt.tight_layout()
        # plt.show()















        


##############################!@#$%%^#############$$$$$$$$$$$$$$$$$$$$$$$
        import xml.etree.ElementTree as ET

        # Function to parse hydraulic conductivity data from XML file
        def parse_hydraulic_conductivity(xml_file):
            #tree = ET.parse(xml_file)
            root = tree.getroot()
            hc_functions = {}

            # Find all hydraulic conductivity functions
            for kfn in root.findall('.//KFn'):
                name = kfn.find('Name').text
                points = []
                for point in kfn.find('Points').findall('Point'):
                    x = float(point.get('X'))
                    y = float(point.get('Y'))
                    points.append((x, y))
                hc_functions[name] = points

            return hc_functions, tree

        # Function to update hydraulic conductivity points in XML
        def update_hydraulic_conductivity_points(kfn, new_points):
            points_element = kfn.find('Points')
            points_element.clear()
            points_element.set('Len', str(len(new_points)))
            for point in new_points:
                point_element = ET.Element('Point')
                point_element.set('X', str(point[0]))
                point_element.set('Y', str(point[1]))
                points_element.append(point_element)

        # Function to parse volumetric water content functions
        def parse_volumetric_water_content(tree):
            root = tree.getroot()
            vwc_functions = {}

            # Find all VolWCFn entries
            for vwc_fn in root.findall('.//VolWCFn'):
                name = vwc_fn.find('Name').text
                points = []
                points_element = vwc_fn.find('Points')
                if points_element is not None:
                    for point in points_element.findall('Point'):
                        x = float(point.get('X'))
                        y = float(point.get('Y'))
                        points.append((x, y))
                vwc_functions[name] = points

            return vwc_functions

        # Function to update volumetric water content points in XML
        def update_volwc_points(tree, function_name, new_points):
            for vwc_fn in tree.findall('.//VolWCFn'):
                name = vwc_fn.find('Name').text
                if name == function_name:
                    points_element = vwc_fn.find('Points')
                    if points_element is not None:
                        points_element.clear()
                        points_element.set('Len', str(len(new_points)))
                        for point in new_points:
                            point_element = ET.Element('Point')
                            point_element.set('X', str(point[0]))
                            point_element.set('Y', str(point[1]))
                            points_element.append(point_element)
                    break

        # Path to XML file
        

        # Load and parse hydraulic conductivity functions
        hydraulic_conductivity_functions, tree = parse_hydraulic_conductivity(xml_file_path)

    

        # Update hydraulic conductivity functions by name
        for kfn in tree.findall('.//KFn'):
            name = kfn.find('Name').text
            if name == 'hc_for_dyke ash':
                update_hydraulic_conductivity_points(kfn, formatted_data_dyke_hc)
            elif name == 'hc_for_pond ash':
                update_hydraulic_conductivity_points(kfn, formatted_data_pond_hc)
            elif name == 'hc_for_foundation':
                update_hydraulic_conductivity_points(kfn, formatted_data_found_hc)
            elif name == 'hc for starter dyke( above)':
                update_hydraulic_conductivity_points(kfn, formatted_data_starter_hc)
  

        # Update VolWCFn points in XML
        #update_volwc_points(tree, 'vwc for starter dyke', new_points_starter_dyke)
        update_volwc_points(tree, 'vwc_for_pond ash', formatted_data_pond_vwc)
        update_volwc_points(tree, 'vwc_for_foundation', formatted_data_found_vwc)
        update_volwc_points(tree, 'vwc_for_dyke ash', formatted_data_dyke_vwc)
        update_volwc_points(tree, 'vwc for starter dyke', formatted_data_starter_vwc)
        # Save the updated XML file
        tree.write(xml_file_path)
##############################$$$$$$$$$$$$$$$$$$%%%%%%%%%%%%%%%%%%%

        # import xml.etree.ElementTree as ET

        # def parse_hydraulic_conductivity(xml_file):
        #     # Parse the XML file
        #     #tree = ET.parse(xml_file)
        #    # root = tree.getroot()

        #     # Dictionary to store hydraulic conductivity functions and their points
        #     hc_functions = {}

        #     # Find all hydraulic conductivity functions
        #     for kfn in root.findall('.//KFn'):
        #         name = kfn.find('Name').text
        #         points = []

        #         # Collect points for this function
        #         for point in kfn.find('Points').findall('Point'):
        #             x = float(point.get('X'))
        #             y = float(point.get('Y'))
        #             points.append((x,y))

        #         # Store the points in the dictionary
        #         hc_functions[name] = points

        #     return hc_functions, tree

        # def update_hydraulic_conductivity_points(kfn, new_points):
        #     # Clear existing points
        #     points_element = kfn.find('Points')
        #     points_element.clear()
        #     # Set the Len attribute to the number of points in new_points
        #     points_element.set('Len', str(len(new_points)))
        #     # Add new points
        #     for point in new_points:
        #         point_element = ET.Element('Point')
        #         point_element.set('X', str(point[0]))
        #         point_element.set('Y', str(point[1]))
        #         points_element.append(point_element)




        # # Path to your XML file
        # xml_file_path1 = "C:/Users/abhia/OneDrive/Documents/abhitest2/acess_xml/for_12m.xml"
        # #print(file_names1[idx])
        # hydraulic_conductivity_functions, tree = parse_hydraulic_conductivity(xml_file_path1)

   

        # # Update points for hc_dyke_ash and hc_pond_ash separately
        # for kfn in tree.findall('.//KFn'):
        #     name = kfn.find('Name').text
        #     if name == 'hc_for_dyke ash':
        #         update_hydraulic_conductivity_points(kfn, data_points_dyke)
        #     elif name == 'hc_for_pond ash':
        #         update_hydraulic_conductivity_points(kfn, data_points_pond)
        #     elif name == 'hc_for_foundation':
        #         update_hydraulic_conductivity_points(kfn, data_points_found)
        #     # elif name == 'hc_for_foundation':
        #     #     update_hydraulic_conductivity_points(kfn, data_points1)   
        # # Write the updated XML back to the file
        # #tree.write(xml_file_path, encoding='utf-8', xml_declaration=True)

        # # Printing the updated points for verification
        # print("Updated Points for hc dyke ash:", data_points_dyke,"\n")
        # print("Updated Points for hc pond ash:", data_points_pond,"\n")
        # print("Updated Points for hc foundation:", data_points_found,"\n")
        
        # print(file_names1[idx])







































        # print(f'Updated dyke ash values: {unitweight3.text}, {cohesion3.text}, {angle3.text}')
        # #material4
        
        # unitweight4.text = str(rows_4.iloc[idx]['Unit Weight (kN/m³)'])
        # cohesion4.text = str(rows_4.iloc[idx]['Cohesion (kPa)'])
        # angle4.text = str(rows_4.iloc[idx]['Angle of Internal Friction (°)'])

        # # permeability_pond_Ash.set("Y", str(rows_4.iloc[idx]['permeability']))
        # print(f'Updated pond ash values: {unitweight4.text}, {cohesion4.text}, {angle4.text}')
        
        # #material5
        # unitweight5.text = str(rows_5.iloc[idx]['Unit Weight (kN/m³)'])
        # cohesion5.text = str(rows_5.iloc[idx]['Cohesion (kPa)'])
        # angle5.text = str(rows_5.iloc[idx]['Angle of Internal Friction (°)'])
        
        # print(f'Updated foundation values: {unitweight5.text}, {cohesion5.text}, {angle5.text}')

        # unitweight6.text = str(rows_6.iloc[idx]['Unit Weight (kN/m³)'])
        # cohesion6.text = str(rows_6.iloc[idx]['Cohesion (kPa)'])
        # angle6.text = str(rows_6.iloc[idx]['Angle of Internal Friction (°)'])
        # print(f'Updated starter dyke (below phreatic line) values: {unitweight6.text}, {cohesion6.text}, {angle6.text}')
        
        
        print("end")
    # Write the updated XML to the corresponding file
    #j=j+1
    # if idx < len(file_names1):
    #     #tree.write(file_names1[idx])
    #     tree.write(file_names1[idx])

    if idx < len(file_names1):
        tree.write(file_names1[idx])
        # tree.write(file_names1[idx])
        print(file_names1[idx])
        
        
        


#         #till here xml files generated successfully


import os
import subprocess

# Define the paths
xml_directory = r"D:\Abinai_Mtech\xmlfiles"  # Directory containing your XML files
geoCmd_path = r"C:\Program Files\Seequent\GeoStudio 11.4\Bin\geoCmd.exe"  # Path to geoCmd.exe

# List all XML files in the directory
xml_files = [file for file in os.listdir(xml_directory) if file.endswith('.xml')]

# Iterate over each XML file
for file in xml_files:
    xml_file_path = os.path.join(xml_directory, file)
    
    print(f"Processing {file}...")

    # Command 1: Solve
    command1 = [geoCmd_path, xml_file_path, "/solve"]
    print(f"Running command1: {command1}")

    try:
        result1 = subprocess.run(command1)
        print(f"Command 1 output: is good ")
    except subprocess.CalledProcessError as e:
        print(f"Error running command1: {e.stderr}")  # Here you should see the error message if there's any failure

    # Command 2: Solve and Generate Report
    command2 = [geoCmd_path, xml_file_path, "/solve", "/report"]
    print(f"Running command2: {command2}")

    try:
        result2 = subprocess.run(command2, check=True, capture_output=True, text=True, cwd=xml_directory)
        print(f"Command 2 output:\n{result2.stdout}")
    except subprocess.CalledProcessError as e:
        print(f"Error running command2: {e.stderr}")

    # Command 3: Solve, Report, and Generate Results
    command3 = [geoCmd_path, xml_file_path, "/solve", "/report", "/results"]
    print(f"Running command3: {command3}")

    try:
        result3 = subprocess.run(command3, check=True, capture_output=True, text=True, cwd=xml_directory)
        print(f"Command 3 output:\n{result3.stdout}")
    except subprocess.CalledProcessError as e:
        print(f"Error running command3: {e.stderr}")

    print(f"Completed processing {file}\n")

######################################################################################################################################################
#############################################end######################################################


# import os
# from bs4 import BeautifulSoup

# # Directory containing the HTML files
# html_directory_path = "C:/Users/abhia/OneDrive/Documents/abhitest2/xmlfiles/"

# # Function to extract saturated Kx values from a single HTML file
# def extract_saturated_kx_values(html_file_path):
#     # Read the HTML file
#     with open(html_file_path, 'r', encoding='utf-8') as file:
#         content = file.read()

#     # Parse the HTML content
#     soup = BeautifulSoup(content, 'html.parser')

#     # List to store the relevant 'hc' entries for this file
#     results = []

#     # Loop through all <h2> tags
#     for h2 in soup.find_all('h2'):
#         title = h2.get_text(strip=True)  # Get the text content of the <h2> tag

#         # Initialize a flag to check if Kx value was found
#         kx_found = False

#         # Find all <p> tags with class 'Report1' within the same parent container
#         for report in h2.find_all_next('p', class_='Report1'):
#             # Check if the current paragraph contains 'Saturated Kx'
#             if 'Saturated Kx' in report.get_text():
#                 # Extract the Kx value
#                 kx_value = report.find('span', class_='Value')
#                 if kx_value:
#                     # Only add results related to 'hc'
#                     if 'hc' in title.lower():
#                         results.append(f"For {title}, the saturated Kx value is: {kx_value.get_text(strip=True)}")
#                         kx_found = True
#                 break  # Stop searching after finding the first match
        
#         # If no Kx value was found, append the message
#         if not kx_found:
#             results.append(f"For {title}, no saturated Kx value found.")

#     # Filter and return only the 'hc' entries
#     return [result for result in results if 'hc' in result.lower()]

# # Function to process all HTML files and store results in a dictionary
# def process_all_html_files(directory_path):
#     # Dictionary to hold the results for each file
#     results_dict = {}

#     for filename in os.listdir(directory_path):
#         if filename.endswith('.html'):
#             file_path = os.path.join(directory_path, filename)
#             print(f"\nProcessing file: {filename}")
#             # Store results in the dictionary with filename as the key
#             results_dict[filename] = extract_saturated_kx_values(file_path)

#     return results_dict

# # Run the function for all files in the directory and store results
# file_results = process_all_html_files(html_directory_path)

# # Print the results_dict to display extracted data for each file
# print("\nComplete Results Dictionary:")
# print(file_results)














# from bs4 import BeautifulSoup


# html_directory = r'C:\Users\abhia\OneDrive\Documents\abhitest2\output' # this is the path for html_directory where the html files are stored(should be changed)
# results = []










# # import os
# # from bs4 import BeautifulSoup

# # # Directory containing the HTML files
# # html_directory_path = "C:/Users/abhia/OneDrive/Documents/abhitest2/xmlfiles/"

# # # Function to extract saturated Kx values from a single HTML file
# # def extract_saturated_kx_values(html_file_path):
# #     # Read the HTML file
# #     with open(html_file_path, 'r', encoding='utf-8') as file:
# #         content = file.read()

# #     # Parse the HTML content
# #     soup = BeautifulSoup(content, 'html.parser')

# #     # List to store the relevant 'hc' entries for this file
# #     results = []

# #     # Loop through all <h2> tags
# #     for h2 in soup.find_all('h2'):
# #         title = h2.get_text(strip=True)  # Get the text content of the <h2> tag

# #         # Initialize a flag to check if Kx value was found
# #         kx_found = False

# #         # Find all <p> tags with class 'Report1' within the same parent container
# #         for report in h2.find_all_next('p', class_='Report1'):
# #             # Check if the current paragraph contains 'Saturated Kx'
# #             if 'Saturated Kx' in report.get_text():
# #                 # Extract the Kx value
# #                 kx_value = report.find('span', class_='Value')
# #                 if kx_value:
# #                     # Only add results related to 'hc'
# #                     if 'hc' in title.lower():
# #                         results.append(f"For {title}, the saturated Kx value is: {kx_value.get_text(strip=True)}")
# #                         kx_found = True
# #                 break  # Stop searching after finding the first match
        
# #         # If no Kx value was found, append the message
# #         if not kx_found:
# #             results.append(f"For {title}, no saturated Kx value found.")

# #     # Filter and return only the 'hc' entries
# #     return [result for result in results if 'hc' in result.lower()]

# # # Function to process all HTML files and store results in a dictionary
# # def process_all_html_files(directory_path):
# #     # Dictionary to hold the results for each file
# #     results_dict = {}

# #     for filename in os.listdir(directory_path):
# #         if filename.endswith('.html'):
# #             file_path = os.path.join(directory_path, filename)
# #             print(f"\nProcessing file: {filename}")
# #             # Store results in the dictionary with filename as the key
# #             results_dict[filename] = extract_saturated_kx_values(file_path)

# #     return results_dict

# # # Run the function for all files in the directory and store results
# # file_results = process_all_html_files(html_directory_path)

# # # Print the results_dict to display extracted data for each file
# # print("\nComplete Results Dictionary:")
# # print(file_results)





# #####################@@@@@@@@@@@@@@@@@!!!!!!!!!!!!!!!!!!!@@@@@@@@@@@#################
# #for file_name in os.listdir(html_directory):
# for idx, file_name in enumerate(os.listdir(html_directory)):
#     output_file = os.path.join(html_directory,file_name) #this will join htmldrectory and file_name example "C:\Users\abhia\OneDrive\Documents\phyton foundation course\comming4\output\file_841 - Slope Stability.html"
#     print(output_file)
#     if(not output_file.endswith('(3).html')):
#        continue
#     with open(output_file, 'r', encoding='utf-8') as file:
#         html_content = file.read()

#     soup = BeautifulSoup(html_content, '.html.parser')
#     factor_of_safety = soup.select_one("p:contains('Factor of Safety') span.Value").text      #finds the p paragraph and fine the factor of safefy and value

#     # Find the <h1> tag containing 'Materials'
#     materials_header = soup.find('h1', text='Materials')

#     # Initialize an empty dictionary to store the results
#     report_data = {}

#     report_data["Factor of Safety"] = float(factor_of_safety)
    
 
        

    

    
#     # for i in range(len(file_results)):
#     #     report_data["Factor of Safety"] = str(factor_of_safety)
#     #     # Using f-string to format the key name correctly
#     #     report_data["permeability"] = file_results[f'file_84{i} - Steady-State Seepage.html']
    



#         # Convert file_results to a list of (filename, content) tuples for easy indexing
#     file_results_list = list(file_results.items())
#     len(file_results_list)
#     #print(file_results_list)
#     # Access the second file (index 1)


#     # Convert dictionary to a list of (key, value) tuples for easy indexing


#     # Iterate over each index
   
#     if idx < len(file_results_list):  # Make sure to check the bounds
#         filename, content_list = file_results_list[idx]  # Access content based on the index

#         print(f"Content for {filename}:")
        
#         content = {}

#         # Print each line in the content_list
#         for line in content_list:
#             key,val = line.split(',')
#             content[key] = val

#         # Store the content list for the current file in report_data
#         report_data["Premeability"] = content
#         report_data["a_found"]= aRes2[idx]
#         report_data["a_pond"]= aRes4[idx]
#         report_data["a_dyke"]= aRes3[idx]
#         report_data["a_starter"]= aRes1[idx]

#         report_data["n_found"]=nRes2[idx]
#         report_data["n_pond"]=nRes4[idx]
#         report_data["n_dyke"]=nRes3[idx]
#         report_data["n_starter"]= nRes1[idx]

#         report_data["m_found"]= mRes2[idx]
#         report_data["m_pond"]= mRes4[idx]
#         report_data["m_dyke"]= mRes3[idx]
#         report_data["m_starter"]= mRes1[idx]

#         report_data["ts_found"]= tsRes2[idx]
#         report_data["ts_pond"]= tsRes4[idx]
#         report_data["ts_dyke"]= tsRes3[idx]
#         report_data["ts_starter"]= tsRes1[idx]

#         report_data["tr_found"]= thRes2[idx]
#         report_data["tr_pond"]= thRes4[idx]
#         report_data["tr_dyke"]= thRes3[idx]
#         report_data["tr_starter"]= thRes1[idx]

#         print('*******************************************')

#         # print(f"Stored data in report_data['{filename}']: {report_data[filename]}\n")








#     # # Outer loop to process each file in the directory (example loop for demonstration)
#     # for idx, file_name in enumerate(os.listdir(html_directory)):
#     #     output_file = os.path.join(html_directory, file_name)
        
#     #     # Skip if it's not an HTML file
#     #     if not output_file.endswith('.html'):
#     #         continue

#     #     # Read the content of the HTML file (assuming some processing here)
#     #     with open(output_file, 'r', encoding='utf-8') as file:
#     #         html_content = file.read()


#     #     file_results_list = list(file_results.items())
#     #     print(file_results_list)

#     #     # Access only the first result from file_results for each iteration
#     #     if idx < len(file_results_list):
#     #         # Retrieve the first (filename, content_list) pair based on the current outer loop index
#     #         single_filename, single_content_list = file_results_list[idx]
            
#     #         # Append this single entry to the report_data
#     #         report_data[single_filename] = single_content_list
            
#     #         # Display content for debugging
#     #         print(f"Content for {single_filename}:")
#     #         for line in single_content_list:
#     #             print(line)
#     #         print("\n")  # Separate output for readability

#     # # report_data now contains only the first file's content for each outer loop iteration
#     # print("Final Report Data:", report_data)


#     #report_data["Factor of Safety"] = str(factor_of_safety)
    
#     if materials_header:
#         # Iterate over siblings of the found <h1> tag
#         for sibling in materials_header.find_next_siblings():
#             if sibling.name == 'h1':
#                 break  # Stop if we encounter another <h1> tag
            
#             if sibling.name == 'h2':
#                 material = sibling.get_text().strip()
#                 material_data = {}
                
#                 # Extract all <p> tags following the <h2> until the next <h2>
#                 for p_tag in sibling.find_next_siblings():
#                     if p_tag.name == 'h2':
#                         break  # Stop at the next <h2> tag
                    
#                     if p_tag.name == 'p':
#                         text = p_tag.get_text()
#                         #print(text)
#                         # test = a,b = [unit text, 463]
#                         # ' asodf   ' = 'aaa'
#                         if ':' in text:
#                             key, value = text.split(':', 1)
#                             material_data[key.strip()] = value.strip().split(' ')[0]
                
#                 # Store this material's data in the main dictionary
#                 report_data[material] = material_data
#                 #print(report_data)
#     results.append(report_data)
# print('abhi')
# print(results)


# output_data = [
   
# ]
# custom_columns = pd.MultiIndex.from_tuples([
#     ('Effective Cohesion(KPA)', 'starter dyke soil (above phreatic line)'),
#     # ('Effective Cohesion(KPA)', 'soil cover'),
#     #('Effective Cohesion(KPA)', 'dyke ash'),
#     ('Effective Cohesion(KPA)', 'pond ash'),
#     ('Effective Cohesion(KPA)', 'foundation'),

#     ('Unit Weight', 'starter dyke soil (above phreatic line)'),
#     # ('Unit Weight', 'soil cover'),
#     #('Unit Weight', 'dyke ash'),
#     ('Unit Weight', 'pond ash'),
#     ('Unit Weight', 'foundation'),
    
#     ('Effective Friction Angle', 'starter dyke soil (above phreatic line)'),
#     # ('Effective Friction Angle', 'soil cover'),
#     #('Effective Friction Angle', 'dyke ash'),
#     ('Effective Friction Angle', 'pond ash'),
#     ('Effective Friction Angle', 'foundation'),

#     ('Premeability', 'Pond ash'),
#     ('Premeability', 'fillter material'),
#     ('Premeability', 'fine aggregate'),
#     ('Premeability', 'course'),
#     ('Premeability', 'starter dyke( above)'),
#     ('Premeability', 'foundation'),
#     #('permeability','dyke ash'),
#     ('a', 'starter dyke'),
#     ('a', 'foundation'),
#     ('a', 'dye ash'),
#     ('a', 'Pond ash'),
    
#     ('n', 'starter dyke'),
#     ('n', 'foundation'),
#     ('n', 'dye ash'),
#     ('n', 'Pond ash'),


#     ('m', 'starter dyke'),
#     ('m', 'foundation'),
#     ('m', 'dye ash'),
#     ('m', 'Pond ash'),
    
    



#     ('theta_s', 'starter dyke'),
#     ('theta_s', 'foundation'),
#     ('theta_s', 'dye ash'),
#     ('theta_s', 'Pond ash'),
    

#     ('theta_r', 'starter dyke'),
#     ('theta_r', 'foundation'),
#     ('theta_r', 'dye ash'),
#     ('theta_r', 'Pond ash'),
    

  


#     ('Factor of Safety', 'Factor of Safety'),

    
# ])

# for item in results:
#     output = [
#         # item['Factor of Safety'],  # Name column left empty
#         item["starter dyke soil (above phreatic line)"]["Effective Cohesion"],  # Effective Cohesion - starter dyke soil
#         # item["soil cover"]["Effective Cohesion"],  # Effective Cohesion - soil cover
#         #item["dyke ash"]["Effective Cohesion"],  # Effective Cohesion - soil cover
#         item["pond ash"]["Effective Cohesion"],  # Effective Cohesion - soil cover
#         item["foundation"]["Effective Cohesion"],  # Effective Cohesion - soil cover

#         item["starter dyke soil (above phreatic line)"]["Unit Weight"],  # Unit Weight - starter dyke soil
#         # item["soil cover"]["Unit Weight"],  # Unit Weight - soil cover
#         #item["dyke ash"]["Unit Weight"],  # Unit Weight - soil cover
#         item["pond ash"]["Unit Weight"],  # Unit Weight - soil cover
#         item["foundation"]["Unit Weight"],  # Unit Weight - soil cover

#         item["starter dyke soil (above phreatic line)"]["Effective Friction Angle"],  # Effective Friction Angle - starter dyke soil
#         # item["soil cover"]["Effective Friction Angle"],  # Effective Friction Angle - soil cover
#         #item["dyke ash"]["Effective Friction Angle"],  # Effective Friction Angle - soil cover
#         item["pond ash"]["Effective Friction Angle"],  # Effective Friction Angle - soil cover
#         item["foundation"]["Effective Friction Angle"],  # Effective Friction Angle - soil cove

#         item["Premeability"]["For hc_for_pond ash"].split(':')[1].strip().split(" ")[0], 
#         item["Premeability"]["For hc for fillter material"].split(':')[1].strip().split(" ")[0], 
#         item["Premeability"]["For hc for fine aggregate"].split(':')[1].strip().split(" ")[0], 
#         item["Premeability"]["For hc for course"].split(':')[1].strip().split(" ")[0], 
#         item["Premeability"]["For hc for starter dyke( above)"].split(':')[1].strip().split(" ")[0], 
#         item["Premeability"]["For hc_for_foundation"].split(':')[1].strip().split(" ")[0], 
 

#         # item[a[item]],aRes[1],aRes[2],
        
        
        
#         item["a_starter"],
#         item["a_found"],
#         item["a_dyke"],
#         item["a_pond"],

#         item["n_starter"],
#         item["n_found"],
#         item["n_dyke"],
#         item["n_pond"],

#         item["m_starter"],
#         item["m_found"],
#         item["m_dyke"],
#         item["m_pond"],


#         item["ts_starter"],
#         item["ts_found"],
#         item["ts_dyke"],
#         item["ts_pond"],

        
        
#         item["tr_starter"],
#         item["tr_found"],
#         item["tr_dyke"],
#         item["tr_pond"],
        
        





       
#         #item["dyke ash"]["permeability"],
#         item['Factor of Safety'] # Name column left empty

#     ]
#     output_data.append(output)


# print(output_data)
# df = pd.DataFrame(output_data, columns=custom_columns)                  # Convert the dictionary to a pandas DataFram
# print(df)
# # df['Factor of Safety'] = pd.to_numeric(df['Factor of Safety'])   #convert to numeric value     
# #  Save the DataFrame to an Excel file
# output_path = os.path.join(html_directory,'output22.xlsx')    #it will create the path (change the excel name accoudingly because same name excel file cannot open)
# df.to_excel(output_path, index=True)                        #it will asssign to that path

# # Confirmation message
# print(f" saved to 'output22.xlsx'.")      #(change this also for understanding user)
######################################################################################################################################################
#############################################end######################################################
from bs4 import BeautifulSoup
import re
import os
import pandas as pd

# Function to extract Material Data from the Slope Stability file
def extract_material_data(html_file_path):
    with open(html_file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    materials_data = {}

    # Extract Factor of Safety manually
    factor_of_safety = None
    for p_tag in soup.find_all('p'):
        if "Factor of Safety" in p_tag.get_text(strip=True):
            factor_of_safety_tag = p_tag.find('span', class_='Value')
            if factor_of_safety_tag:
                factor_of_safety = factor_of_safety_tag.get_text(strip=True)
                break

    if factor_of_safety:
        materials_data["Factor of Safety"] = factor_of_safety

    radius = None
    for p_tag in soup.find_all('p'):
        if "Radius:" in p_tag.get_text(strip=True):
            radius_tag = p_tag.find('span', class_='Value')
            if radius_tag:
                radius = radius_tag.get_text(strip=True)
                break

    if radius:
        materials_data["Radius"] = radius

    center = None
    for p_tag in soup.find_all('p'):
        if "Center" in p_tag.get_text(strip=True):
            center_tag = p_tag.find('span', class_='Value')
            if center_tag:
                center = center_tag.get_text(strip=True)
                break

    if center:
        materials_data["Center"] = center


        

    # Find the 'Materials' section
    materials_header = soup.find('h1', string='Materials') or soup.find('h2', string='Materials')
    
    if materials_header:
        for sibling in materials_header.find_next_siblings():
            if sibling.name == 'h1':
                break
            
            if sibling.name == 'h2':
                material_name = sibling.get_text(strip=True)
                material_data = {}

                for p_tag in sibling.find_next_siblings():
                    if p_tag.name == 'h2':
                        break
                    if p_tag.name == 'p':
                        text = p_tag.get_text(strip=True)
                        if ':' in text:
                            key, value = text.split(':', 1)
                            material_data[key.strip()] = value.strip()
                
                if material_data:
                    materials_data[material_name] = material_data

    return materials_data

# Function to extract Saturated Kx values and add to Materials section
def extract_saturated_kx_values(html_file_path, materials_data):
    with open(html_file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')

    for h2 in soup.find_all('h2'):
        title = h2.get_text(strip=True)

        for report in h2.find_all_next('p', class_='Report1'):
            if 'Saturated Kx' in report.get_text():
                kx_value = report.find('span', class_='Value')
                if kx_value and 'hc' in title.lower():
                    materials_data[title] = f"Saturated Kx value: {kx_value.get_text(strip=True)}"
                break

    return materials_data

# Main function to organize extracted data
def collect_geostudio_data(source_directory):
    combined_data = {}
    
    for filename in os.listdir(source_directory):
        file_path = os.path.join(source_directory, filename)
        base_name = re.match(r"(file_\d+)", filename)
        if base_name:
            base_name = base_name.group(1)

            if base_name not in combined_data:
                combined_data[base_name] = {}

            if "Slope Stability" in filename:
                combined_data[base_name]["Materials"] = extract_material_data(file_path)

            elif "Steady-State Seepage" in filename:
                if "Materials" not in combined_data[base_name]:
                    combined_data[base_name]["Materials"] = {}
                combined_data[base_name]["Materials"] = extract_saturated_kx_values(file_path, combined_data[base_name]["Materials"])

    return combined_data


# Define the source directory
source_directory = r"D:\Abinai_Mtech\xmlfiles"

# Run the data collection and print the results
geostudio_data = collect_geostudio_data(source_directory)
for file, data in geostudio_data.items():
    print(f"\nFilename: {file}")
    for key, value in data.items():
        print(f"{key}: {value}")

print("##################")        
print(geostudio_data)
print("#####################")
# print(geostudio_data[0])




# Custom column structure (MultiIndex)
custom_columns = pd.MultiIndex.from_tuples([
    ('filename','file name'),
    ('Effective Cohesion(KPA)', 'starter dyke soil'),
    ('Effective Cohesion(KPA)', 'pond ash'),
    ('Effective Cohesion(KPA)', 'foundation'),
    ('Effective Cohesion(KPA)', 'dyke ash'),
    ('Unit Weight', 'starter dyke soil'),
    ('Unit Weight', 'pond ash'),
    ('Unit Weight', 'foundation'),
    ('Unit Weight', 'dyke ash'),
    ('Effective Friction Angle', 'starter dyke soil'),
    ('Effective Friction Angle', 'pond ash'),
    ('Effective Friction Angle', 'foundation'),
    ('Effective Friction Angle', 'dyke ash'),
    ('Premeability(m/s)', 'Pond ash'),
    ('Premeability(m/s)', 'fillter material'),
    ('Premeability(m/s)', 'fine aggregate'),
    ('Premeability(m/s)', 'course'),
    ('Premeability(m/s)', 'starter dyke'),
    ('Premeability(m/s)', 'foundation'),
    ('Premeability(m/s)', 'dyke ash'),
    ('Factor of Safety', 'Factor of Safety'),
    ('Radius(m)', 'Radius'),
    ('Center(m)', 'Center'),




    
])
# Process the results and extract values
output_data = []

for file_name, data in geostudio_data.items():
    materials = data['Materials']
    
    # Extracting values from the nested structure
    output = [
        file_name,
        materials.get("starter dyke soil", {}).get("Effective Cohesion", "N/A").split(' ')[0],
        materials.get("pond ash", {}).get("Effective Cohesion", "N/A").split(' ')[0],
        materials.get("foundation", {}).get("Effective Cohesion", "N/A").split(' ')[0],

        materials.get("dyke ash", {}).get("Effective Cohesion", "N/A").split(' ')[0],
            
        materials.get("starter dyke soil", {}).get("Unit Weight", "N/A").split(' ')[0],
        materials.get("pond ash", {}).get("Unit Weight", "N/A").split(' ')[0],
        materials.get("foundation", {}).get("Unit Weight", "N/A").split(' ')[0],
         
        
        materials.get("dyke ash", {}).get("Unit Weight", "N/A").split(' ')[0],
                   
        materials.get("starter dyke soil", {}).get("Effective Friction Angle", "N/A").split(' ')[0],
        materials.get("pond ash", {}).get("Effective Friction Angle", "N/A").split(' ')[0],
        materials.get("foundation", {}).get("Effective Friction Angle", "N/A").split(' ')[0],
        

        materials.get("dyke ash", {}).get("Effective Friction Angle", "N/A").split(' ')[0],

        # Handle permeability values safely
        materials.get("hc_for_pond ash", {}).split(':')[1].strip().split()[0],
        materials.get("hc for fillter material", {}).split(':')[1].strip().split()[0],
        materials.get("hc for fine aggregate", {}).split(':')[1].strip().split()[0],
        materials.get("hc for course", {}).split(':')[1].strip().split()[0],
        materials.get("hc for starter dyke( above)", {}).split(':')[1].strip().split()[0],
        materials.get("hc_for_foundation", {}).split(':')[1].strip().split()[0],

        materials.get("hc_for_dyke ash", {}).split(':')[1].strip().split()[0],
        #materials.get("hc_for_foundation", {}).split(':')[1].strip(),
        #materials.get["Saturated Kx Values"][0].split("the saturated Kx value is:")[1].strip().split()[0],
        # materials.get("Saturated Kx Values", {}).get("For hc for fillter material", "N/A").split(':')[0].strip() ,
        # materials.get("Saturated Kx Values", {}).get("For hc for fine aggregate", "N/A").split(':')[0].strip() ,
        # materials.get("Saturated Kx Values", {}).get("For hc for course", "N/A").split(':')[0].strip() ,
        # materials.get("Saturated Kx Values", {}).get("For hc for starter dyke( above)", "N/A").split(':')[0].strip() ,
        # materials.get("Saturated Kx Values", {}).get("For hc_for_foundation", "N/A").split(':')[0].strip() ,
        
        materials.get('Factor of Safety', 'N/A'),
        materials.get('Radius', 'N/A'),
        materials.get('Center', 'N/A')


    ]
    output_data.append(output)
# Convert the output data into a DataFrame
df = pd.DataFrame(output_data, columns=custom_columns)
print(df)
# Save the DataFrame to an Excel file
output_path = os.path.join('D:/Abinai_Mtech/xmlfiles', 'output22.xlsx')
df.to_excel(output_path, index=True)

# Confirmation message
print(f"Data has been saved to '{output_path}'.")





#################### this part of code will arrange the excel file line by line #####################

'''import pandas as pd
import openpyxl
import re

# Load the Excel file using openpyxl (to preserve merged cells)
file_path = 'C:/Users/abhia/OneDrive/Documents/abhitest2/xmlfiles/output22.xlsx'

# Open the workbook and select the sheet using openpyxl
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Extract the merged cells in the first two rows
merged_cells = sheet.merged_cells.ranges

# Read the data into a DataFrame using pandas
df = pd.read_excel(file_path, engine='openpyxl', header=None)

# Extract the first two rows (which have merged cells)
first_two_rows = df.iloc[:2]

# Extract the remaining rows (which will be sorted)
remaining_rows = df.iloc[2:]

# Function to extract the numerical part from filenames
def extract_number(filename):
    if isinstance(filename, str):  # Only process if it's a string
        match = re.search(r'file_(\d+)', filename)
        if match:
            return int(match.group(1))  # Extract the number from "file_X"
    return float('inf')  # If no match or not a string, return a large number to keep it at the end

# The column containing filenames is assumed to be the second column (index 1)
filename_column_index = 1

# Sort the remaining rows based on the numerical part of filenames
sorted_remaining_rows = remaining_rows.sort_values(by=filename_column_index, key=lambda x: x.apply(extract_number))

# Concatenate the first two rows with the sorted remaining rows
final_df = pd.concat([first_two_rows, sorted_remaining_rows], ignore_index=True)

# Save the result back to Excel
# Create a new workbook and copy the final_df content into it
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Write the sorted data into the new sheet (starting from the first row)
for r_idx, row in enumerate(final_df.values, 1):
    for c_idx, value in enumerate(row, 1):
        new_sheet.cell(row=r_idx, column=c_idx, value=value)

# Reapply the merged cells from the original sheet to the new sheet
for merged_range in merged_cells:
    new_sheet.merge_cells(str(merged_range))

# Save the final output file with merged cells intact
final_file_path = 'C:/Users/abhia/OneDrive/Documents/abhitest2/xmlfiles/sorted_output22.xlsx'
new_workbook.save(final_file_path)

print(f"The file has been sorted with the first two rows unchanged, merged cells preserved, and saved as {final_file_path}.")
 '''
 #####################end of arranging the code line by line ####################

import pandas as pd
import openpyxl
import re

# Load the Excel file using openpyxl (to preserve merged cells)
file_path = 'D:/Abinai_Mtech/xmlfiles/output22.xlsx'

# Open the workbook and select the sheet using openpyxl
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Extract the merged cells in the first two rows
merged_cells = sheet.merged_cells.ranges

# Read the data into a DataFrame using pandas
df = pd.read_excel(file_path, engine='openpyxl', header=None)

# Extract the first two rows (which have merged cells)
first_two_rows = df.iloc[:2]

# Extract the remaining rows (which will be sorted)
remaining_rows = df.iloc[2:]

# Function to extract the numerical part from filenames
def extract_number(filename):
    if isinstance(filename, str):  # Only process if it's a string
        match = re.search(r'file_(\d+)', filename)
        if match:
            return int(match.group(1))  # Extract the number from "file_X"
    return float('inf')  # If no match or not a string, return a large number to keep it at the end

# The column containing filenames is assumed to be the second column (index 1)
filename_column_index = 1

# Sort the remaining rows based on the numerical part of filenames
sorted_remaining_rows = remaining_rows.sort_values(by=filename_column_index, key=lambda x: x.apply(extract_number))

# Concatenate the first two rows with the sorted remaining rows
final_df = pd.concat([first_two_rows, sorted_remaining_rows], ignore_index=True)

# Create a new workbook and copy the final_df content into it
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Write the sorted data into the new sheet (starting from the first row)
for r_idx, row in enumerate(final_df.values, 1):
    for c_idx, value in enumerate(row, 1):
        new_sheet.cell(row=r_idx, column=c_idx, value=value)

# Reapply the merged cells from the original sheet to the new sheet
for merged_range in merged_cells:
    new_sheet.merge_cells(str(merged_range))

# Save the final output file with merged cells intact
final_file_path = 'D:/Abinai_Mtech/xmlfiles/sorted_output22.xlsx'
new_workbook.save(final_file_path)

print(f"The file has been sorted with the first two rows unchanged, merged cells preserved, and saved as {final_file_path}.")





################ end of sortted ###############

import pandas as pd
import xml.etree.ElementTree as ET
import pandas as pd

# Load the XML data
#tree = ET.parse('D:/Abinai_Mtech/acess_xml/for_12m.xml')
root = tree.getroot()

# Define the file names for generated XML files
#filenames1 = [f'C:/Users/abhia/OneDrive/Documents/abhitest2/xmlfiles/file{i}.xml' for i in range(1, 6)]

# Read parameters from Excel


# Identify rows for each material section
material_indices = parameters[parameters['Soil Type'].str.contains('material-', na=False)].index
material_sections = {}

# Define how many values you want to extract for each material section
# num_values = 5  # Example: Get the first 3 rows from each material section, change this as needed

# Store material sections in a dictionary
for i in range(len(material_indices)):
    start_index = material_indices[i] + 1
    end_index = material_indices[i + 1] if i < len(material_indices) - 1 else len(parameters)
    material_rows = parameters.iloc[start_index:end_index]

    # Limit the number of rows based on the 'num_values' parameter
    material_sections[f'material{i + 1}'] = material_rows.head(num_values)

# Access each section separately
rows_1 = material_sections.get('material1', pd.DataFrame())
rows_2 = material_sections.get('material2', pd.DataFrame())
rows_3 = material_sections.get('material3', pd.DataFrame())
rows_4 = material_sections.get('material4', pd.DataFrame())

# Initialize dictionaries to store values of a, m, n for each section
a_values = {'starter dyke soil': [], 'pond ash': [], 'foundation': [], 'dye ash': []}
m_values = {'starter dyke soil': [], 'pond ash': [], 'foundation': [], 'dye ash': []}
n_values = {'starter dyke soil': [], 'pond ash': [], 'foundation': [], 'dye ash': []}
theta_s_values = {'starter dyke soil': [], 'pond ash': [], 'foundation': [], 'dye ash': []}
theta_r_values = {'starter dyke soil': [], 'pond ash': [], 'foundation': [], 'dye ash': []}

# Extract a, m, n, theta_s, and theta_r values from each material section
for idx in range(len(rows_1)):
    # Starter Dyke Soil (above phreatic line)
    a_values['starter dyke soil'].append(rows_1.iloc[idx].get('a', ''))
    m_values['starter dyke soil'].append(rows_1.iloc[idx].get('m', ''))
    n_values['starter dyke soil'].append(rows_1.iloc[idx].get('n', ''))
    theta_s_values['starter dyke soil'].append(0.1)
    theta_r_values['starter dyke soil'].append(0.01)

    # Pond Ash
    a_values['pond ash'].append(rows_4.iloc[idx].get('a', ''))
    m_values['pond ash'].append(rows_4.iloc[idx].get('m', ''))
    n_values['pond ash'].append(rows_4.iloc[idx].get('n', ''))
    theta_s_values['pond ash'].append(0.5)
    theta_r_values['pond ash'].append(0.04)

    # Foundation
    a_values['foundation'].append(rows_2.iloc[idx].get('a', ''))
    m_values['foundation'].append(rows_2.iloc[idx].get('m', ''))
    n_values['foundation'].append(rows_2.iloc[idx].get('n', ''))
    theta_s_values['foundation'].append(0.4)
    theta_r_values['foundation'].append(0.03)

    # Dye Ash (fixed values for all)
    a_values['dye ash'].append(rows_3.iloc[idx].get('a', ''))
    m_values['dye ash'].append(rows_3.iloc[idx].get('m', ''))
    n_values['dye ash'].append(rows_3.iloc[idx].get('n', ''))
    theta_s_values['dye ash'].append(0.3)
    theta_r_values['dye ash'].append(0.02)

# Create a MultiIndex DataFrame with 'a', 'm', 'n', 'theta_s', 'theta_r' as top-level columns and 'starter dyke soil', 'pond ash', 'foundation', 'dye ash' as sub-columns
multi_index_columns = pd.MultiIndex.from_tuples([
    ('a', 'starter dyke soil'),
    ('a', 'pond ash'),
    ('a', 'foundation'),
    ('a', 'dye ash'),
    ('m', 'starter dyke soil'),
    ('m', 'pond ash'),
    ('m', 'foundation'),
    ('m', 'dye ash'),
    ('n', 'starter dyke soil'),
    ('n', 'pond ash'),
    ('n', 'foundation'),
    ('n', 'dye ash'),
    ('theta_s', 'starter dyke soil'),
    ('theta_s', 'pond ash'),
    ('theta_s', 'foundation'),
    ('theta_s', 'dye ash'),
    ('theta_r', 'starter dyke soil'),
    ('theta_r', 'pond ash'),
    ('theta_r', 'foundation'),
    ('theta_r', 'dye ash')
])

output_data = pd.DataFrame({
    ('a', 'starter dyke soil'): a_values['starter dyke soil'],
    ('a', 'pond ash'): a_values['pond ash'],
    ('a', 'foundation'): a_values['foundation'],
    ('a', 'dye ash'): a_values['dye ash'],
    ('m', 'starter dyke soil'): m_values['starter dyke soil'],
    ('m', 'pond ash'): m_values['pond ash'],
    ('m', 'foundation'): m_values['foundation'],
    ('m', 'dye ash'): m_values['dye ash'],
    ('n', 'starter dyke soil'): n_values['starter dyke soil'],
    ('n', 'pond ash'): n_values['pond ash'],
    ('n', 'foundation'): n_values['foundation'],
    ('n', 'dye ash'): n_values['dye ash'],
    ('theta_s', 'starter dyke soil'): theta_s_values['starter dyke soil'],
    ('theta_s', 'pond ash'): theta_s_values['pond ash'],
    ('theta_s', 'foundation'): theta_s_values['foundation'],
    ('theta_s', 'dye ash'): theta_s_values['dye ash'],
    ('theta_r', 'starter dyke soil'): theta_r_values['starter dyke soil'],
    ('theta_r', 'pond ash'): theta_r_values['pond ash'],
    ('theta_r', 'foundation'): theta_r_values['foundation'],
    ('theta_r', 'dye ash'): theta_r_values['dye ash']
}, columns=multi_index_columns)

# Save the DataFrame to an Excel file with MultiIndex columns
output_file = "D:/Abinai_Mtech/xmlfiles/output_values_with_dye_ash.xlsx"
with pd.ExcelWriter(output_file) as writer:
    output_data.to_excel(writer, sheet_name="Sheet1", index=True, header=True)

print("Data extracted and saved to output_values_with_dye_ash.xlsx successfully.")

###################### till here a n m and tr and ts created excel file ########################

######################### removing of blank row in the a n m excel file #############
import pandas as pd
import openpyxl
import re

# Load the Excel file using openpyxl (to preserve merged cells)
file_path = 'D:/Abinai_Mtech/xmlfiles/output_values_with_dye_ash.xlsx'

# Open the workbook and select the sheet using openpyxl
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Extract the merged cells in the first two rows
merged_cells = sheet.merged_cells.ranges

# Read the data into a DataFrame using pandas
df = pd.read_excel(file_path, engine='openpyxl', header=None)

# Extract the first two rows (which have merged cells)
first_two_rows = df.iloc[:2]

# Extract the remaining rows (which will stay as-is, no sorting)
remaining_rows = df.iloc[3:]

# Concatenate the first two rows with the remaining rows (no sorting)
final_df = pd.concat([first_two_rows, remaining_rows], ignore_index=False)

# Create a new workbook and copy the final_df content into it
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Write the data into the new sheet (starting from the first row)
for r_idx, row in enumerate(final_df.values, 1):
    for c_idx, value in enumerate(row, 1):
        new_sheet.cell(row=r_idx, column=c_idx, value=value)

# Reapply the merged cells from the original sheet to the new sheet
for merged_range in merged_cells:
    new_sheet.merge_cells(str(merged_range))

# Save the final output file with merged cells intact
final_file_path = 'D:/Abinai_Mtech/xmlfiles/sorted_output_values_with_dye_ash.xlsx'
new_workbook.save(final_file_path)

print(f"The file has been saved with the first two rows unchanged, merged cells preserved, and saved as {final_file_path}.")

############### end of blank a n m #####################
############## merging of two excel sheets #########################

# import pandas as pd

# # Load the first and second Excel files into pandas DataFrames
# file1_path = 'C:/Users/abhia/OneDrive/Documents/abhitest2/xmlfiles/sorted_output22.xlsx'
# file2_path = 'C:/Users/abhia/OneDrive/Documents/abhitest2/xmlfiles/sorted_output_values_with_dye_ash.xlsx'

# # Read the data into pandas DataFrames
# df1 = pd.read_excel(file1_path, engine='openpyxl', header=0)  # Use header=0 to read column names
# df2 = pd.read_excel(file2_path, engine='openpyxl', header=0)

# # Clean the column names in df1 and df2 by stripping leading/trailing whitespaces
# df1.columns = df1.columns.str.strip()
# df2.columns = df2.columns.str.strip()

# # Check column names in df1
# print("Column names in the first file:")
# print(df1.columns)

# # Define the "Factor of Safety" column name
# factor_of_safety_column_name = 'Factor of Safety'

# # Check if the "Factor of Safety" column exists in df1
# if factor_of_safety_column_name in df1.columns:
#     # Find the index of the "Factor of Safety" column
#     factor_of_safety_column_index = df1.columns.get_loc(factor_of_safety_column_name)
# else:
#     raise ValueError(f"Column '{factor_of_safety_column_name}' not found in the first file")

# # Merge the second file (df2) columns into the first file (df1) before the "Factor of Safety" column
# df1_part_before_safety = df1.iloc[:, :factor_of_safety_column_index]
# df1_part_after_safety = df1.iloc[:, factor_of_safety_column_index:]

# # Concatenate the second file columns before the "Factor of Safety" column
# final_df = pd.concat([df1_part_before_safety, df2, df1_part_after_safety], axis=1)

# # Save the final merged DataFrame into a new Excel file
# final_file_path = 'C:/Users/abhia/OneDrive/Documents/abhitest2/xmlfiles/merged_output.xlsx'
# final_df.to_excel(final_file_path, index=False)

# print(f"Files have been merged and saved as {final_file_path}.")


import openpyxl
import pandas as pd

# Define the paths for the source and target files
source_file = "D:/Abinai_Mtech/xmlfiles/sorted_output22.xlsx"
target_file = "D:/Abinai_Mtech/xmlfiles/sorted_output_values_with_dye_ash.xlsx"

# Load both workbooks using openpyxl
source_wb = openpyxl.load_workbook(source_file)
target_wb = openpyxl.load_workbook(target_file)

# Load the first sheet from both workbooks
source_sheet = source_wb.active  # Or specify the sheet name like 'source_wb['Sheet1']'
target_sheet = target_wb.active  # Or specify the sheet name like 'target_wb['Sheet1']'

# Extract the data from both sheets into pandas DataFrames
source_df = pd.DataFrame(source_sheet.values)
target_df = pd.DataFrame(target_sheet.values)

# Concatenate the two DataFrames side by side
combined_df = pd.concat([source_df, target_df], axis=1)

# Create a new workbook to save the combined data
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Write the combined data to the new sheet
for r_idx, row in enumerate(combined_df.values, 1):
    for c_idx, value in enumerate(row, 1):
        new_sheet.cell(row=r_idx, column=c_idx, value=value)

# Save the new workbook with combined data
output_file = "D:/Abinai_Mtech/xmlfiles/combined_output.xlsx"
new_wb.save(output_file)

print(f"Data from both files have been combined and saved to this folder suceessfully  {output_file}.")
print("successfully code was completed")

